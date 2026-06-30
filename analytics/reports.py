import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from students.models import StudentProfile
from placements.models import Application, PlacementDrive
from companies.models import CompanyProfile
from django.db.models import Count, Q
from django.utils import timezone

def get_placement_statistics():
    """
    Computes placement analytics using Pandas DataFrames.
    """
    students_qs = StudentProfile.objects.all().values('register_number', 'department', 'placement_status', 'cgpa')
    df_students = pd.DataFrame(list(students_qs))

    apps_qs = Application.objects.all().values('student__register_number', 'drive__job_title', 'drive__company__company_name', 'status', 'drive__salary_package', 'applied_at')
    df_apps = pd.DataFrame(list(apps_qs))

    drives_qs = PlacementDrive.objects.all().values('company__company_name', 'job_title', 'salary_package', 'status')
    df_drives = pd.DataFrame(list(drives_qs))

    stats = {
        'total_students': len(df_students),
        'placement_percentage': 0,
        'dept_placement': {},
        'company_hiring': {},
        'highest_package': 0.0,
        'average_package': 0.0,
        'monthly_trends': {},
    }

    if df_students.empty:
        return stats

    # 1. Total Placement %
    placed_mask = df_students['placement_status'].isin(['PLACED', 'DREAM'])
    total_placed = placed_mask.sum()
    stats['placement_percentage'] = int((total_placed / len(df_students)) * 100)

    # 2. Department-wise Placement Rates
    dept_totals = df_students.groupby('department')['register_number'].count()
    dept_placed = df_students[placed_mask].groupby('department')['register_number'].count()
    for dept in df_students['department'].unique():
        total = dept_totals.get(dept, 0)
        placed = dept_placed.get(dept, 0)
        stats['dept_placement'][dept] = {
            'total': int(total),
            'placed': int(placed),
            'rate': int((placed / total * 100)) if total > 0 else 0
        }

    # 3. Company-wise Hiring volumes & Packages
    if not df_apps.empty:
        offered_apps = df_apps[df_apps['status'] == 'OFFERED']
        if not offered_apps.empty:
            hiring_counts = offered_apps.groupby('drive__company__company_name')['student__register_number'].count()
            for company, count in hiring_counts.items():
                stats['company_hiring'][company] = int(count)

            # Packages
            stats['highest_package'] = float(offered_apps['drive__salary_package'].max())
            stats['average_package'] = round(float(offered_apps['drive__salary_package'].mean()), 2)
        else:
            # Fallback if applications exist but none offered yet
            if not df_drives.empty:
                stats['highest_package'] = float(df_drives['salary_package'].max())
                stats['average_package'] = round(float(df_drives['salary_package'].mean()), 2)

        # 4. Monthly Placement Trends (hiring per month)
        offered_apps = df_apps[df_apps['status'] == 'OFFERED'].copy()
        if not offered_apps.empty:
            offered_apps['applied_at'] = pd.to_datetime(offered_apps['applied_at'])
            offered_apps['month'] = offered_apps['applied_at'].dt.strftime('%B %Y')
            monthly = offered_apps.groupby('month')['student__register_number'].count()
            for month, count in monthly.items():
                stats['monthly_trends'][month] = int(count)

    elif not df_drives.empty:
        stats['highest_package'] = float(df_drives['salary_package'].max())
        stats['average_package'] = round(float(df_drives['salary_package'].mean()), 2)

    return stats


def generate_excel_report(report_type, drive_id=None):
    """
    Generates a beautifully styled Excel Workbook using OpenPyXL.
    Types: 'eligible_students', 'selected_students', 'company_reports', 'analytics_reports'
    """
    wb = Workbook()
    ws = wb.active

    # Premium Fonts & Colors (Deep Navy #0F172A, Accent Indigo #6366F1, Borders Slate #E2E8F0)
    font_title = Font(name='Segoe UI', size=16, bold=True, color='FFFFFF')
    font_header = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    font_body = Font(name='Segoe UI', size=10)
    font_bold_body = Font(name='Segoe UI', size=10, bold=True)

    fill_title = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
    fill_header = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
    fill_zebra = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')

    border_thin = Side(border_style="thin", color="CBD5E1")
    border_double = Side(border_style="double", color="0F172A")
    cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    total_border = Border(top=border_thin, bottom=border_double)

    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    if report_type == 'eligible_students' and drive_id:
        drive = PlacementDrive.objects.get(id=drive_id)
        ws.title = "Eligible Students"

        # Title Row
        ws.merge_cells("A1:G1")
        ws['A1'] = f"Eligible Candidates for: {drive.job_title} at {drive.company.company_name}"
        ws['A1'].font = font_title
        ws['A1'].fill = fill_title
        ws['A1'].alignment = align_center
        ws.row_dimensions[1].height = 40

        # Headers
        headers = ["Register No", "Student Name", "Email", "Department", "CGPA", "Backlogs", "Skills"]
        ws.append([]) # Blank spacer
        ws.append(headers)
        
        ws.row_dimensions[3].height = 25
        for col_num in range(1, 8):
            cell = ws.cell(row=3, column=col_num)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center

        # Load students
        students = StudentProfile.objects.all().select_related('user')
        row_idx = 4
        for student in students:
            is_eligible, _ = drive.is_eligible(student)
            if is_eligible:
                row_data = [
                    student.register_number,
                    student.user.get_full_name(),
                    student.user.email,
                    student.get_department_display(),
                    float(student.cgpa),
                    student.backlogs,
                    student.skills
                ]
                ws.append(row_data)
                
                # Apply row styling
                ws.row_dimensions[row_idx].height = 20
                for col_idx in range(1, 8):
                    c = ws.cell(row=row_idx, column=col_idx)
                    c.font = font_body
                    c.border = cell_border
                    if row_idx % 2 == 0:
                        c.fill = fill_zebra
                    
                    if col_idx in [1, 4, 5, 6]:
                        c.alignment = align_center
                    else:
                        c.alignment = align_left
                row_idx += 1

    elif report_type == 'selected_students':
        ws.title = "Selected Students"

        # Title Row
        ws.merge_cells("A1:F1")
        ws['A1'] = "Placement Selection & Offers Report"
        ws['A1'].font = font_title
        ws['A1'].fill = fill_title
        ws['A1'].alignment = align_center
        ws.row_dimensions[1].height = 40

        headers = ["Register No", "Student Name", "Department", "Selected Company", "Designation", "Package (LPA)"]
        ws.append([]) # Spacer
        ws.append(headers)
        ws.row_dimensions[3].height = 25
        for col_num in range(1, 7):
            cell = ws.cell(row=3, column=col_num)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center

        apps = Application.objects.filter(status=Application.Status.OFFERED).select_related('student__user', 'drive__company')
        row_idx = 4
        for app in apps:
            row_data = [
                app.student.register_number,
                app.student.user.get_full_name(),
                app.student.get_department_display(),
                app.drive.company.company_name,
                app.drive.job_title,
                float(app.drive.salary_package)
            ]
            ws.append(row_data)
            
            ws.row_dimensions[row_idx].height = 20
            for col_idx in range(1, 7):
                c = ws.cell(row=row_idx, column=col_idx)
                c.font = font_body
                c.border = cell_border
                if row_idx % 2 == 0:
                    c.fill = fill_zebra
                
                if col_idx in [1, 3, 6]:
                    c.alignment = align_center
                else:
                    c.alignment = align_left
            row_idx += 1

    elif report_type == 'company_reports':
        ws.title = "Company Drives Summary"

        ws.merge_cells("A1:F1")
        ws['A1'] = "Company Drives and Hiring Records"
        ws['A1'].font = font_title
        ws['A1'].fill = fill_title
        ws['A1'].alignment = align_center
        ws.row_dimensions[1].height = 40

        headers = ["Company Name", "Job Title", "Salary Package (LPA)", "Drive Date", "Total Applicants", "Selections"]
        ws.append([])
        ws.append(headers)
        ws.row_dimensions[3].height = 25
        for col_num in range(1, 7):
            cell = ws.cell(row=3, column=col_num)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center

        drives = PlacementDrive.objects.all().select_related('company').annotate(
            applicant_count=Count('applications'),
            placed_count=Count('applications', filter=Q(applications__status=Application.Status.OFFERED))
        )
        row_idx = 4
        for drive in drives:
            row_data = [
                drive.company.company_name,
                drive.job_title,
                float(drive.salary_package),
                drive.drive_date.strftime('%Y-%m-%d'),
                drive.applicant_count,
                drive.placed_count
            ]
            ws.append(row_data)
            
            ws.row_dimensions[row_idx].height = 20
            for col_idx in range(1, 7):
                c = ws.cell(row=row_idx, column=col_idx)
                c.font = font_body
                c.border = cell_border
                if row_idx % 2 == 0:
                    c.fill = fill_zebra
                
                if col_idx in [3, 4, 5, 6]:
                    c.alignment = align_center
                else:
                    c.alignment = align_left
            row_idx += 1

    elif report_type == 'analytics_reports':
        ws.title = "Overview Statistics"
        
        # General overview
        ws.merge_cells("A1:C1")
        ws['A1'] = "Placement Management Portal Analytics Summary"
        ws['A1'].font = font_title
        ws['A1'].fill = fill_title
        ws['A1'].alignment = align_center
        ws.row_dimensions[1].height = 40

        # Load metrics using pandas
        stats = get_placement_statistics()
        
        ws.append([])
        ws.append(["Key Metrics", "Value"])
        ws.cell(row=3, column=1).font = font_header
        ws.cell(row=3, column=1).fill = fill_header
        ws.cell(row=3, column=2).font = font_header
        ws.cell(row=3, column=2).fill = fill_header
        ws.row_dimensions[3].height = 25

        overview_rows = [
            ["Total Students Registered", stats['total_students']],
            ["Placement Percentage (%)", f"{stats['placement_percentage']}%"],
            ["Highest Salary Package Offered (LPA)", f"{stats['highest_package']} LPA"],
            ["Average Salary Package (LPA)", f"{stats['average_package']} LPA"],
        ]

        row_idx = 4
        for metric, val in overview_rows:
            ws.append([metric, val])
            ws.cell(row=row_idx, column=1).font = font_bold_body
            ws.cell(row=row_idx, column=1).border = cell_border
            ws.cell(row=row_idx, column=2).font = font_body
            ws.cell(row=row_idx, column=2).border = cell_border
            ws.cell(row=row_idx, column=2).alignment = align_center
            row_idx += 1

        # Department performance table
        ws.append([])
        row_idx += 1
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
        title_cell = ws.cell(row=row_idx, column=1)
        title_cell.value = "Department-wise Statistics"
        title_cell.font = font_header
        title_cell.fill = fill_header
        title_cell.alignment = align_center
        ws.row_dimensions[row_idx].height = 25
        row_idx += 1

        headers = ["Department", "Registered Students", "Placed Students", "Placement Rate (%)"]
        for col_idx, text in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = text
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
        ws.row_dimensions[row_idx].height = 25
        row_idx += 1

        start_body_row = row_idx
        for dept, info in stats['dept_placement'].items():
            ws.cell(row=row_idx, column=1, value=dept).alignment = align_left
            ws.cell(row=row_idx, column=2, value=info['total']).alignment = align_center
            ws.cell(row=row_idx, column=3, value=info['placed']).alignment = align_center
            ws.cell(row=row_idx, column=4, value=f"{info['rate']}%").alignment = align_center
            
            for col in range(1, 5):
                cell = ws.cell(row=row_idx, column=col)
                cell.font = font_body
                cell.border = cell_border
                if row_idx % 2 == 0:
                    cell.fill = fill_zebra
            row_idx += 1

    # Auto-fit columns to content
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            # Check if cell has coordinate and is part of merged cells
            if hasattr(cell, 'coordinate') and cell.coordinate in ws.merged_cells:
                continue
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    return wb
